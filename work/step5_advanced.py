"""
Step 5: Advanced Analytics → Excel Report
- Reads aggregated/ data from step2 (monthly, cust_agg, prod_agg, sale_full)
- Performs 6 advanced analyses beyond the basic step3 report:

    1. Seasonality Decomposition — split monthly sales into trend + seasonal + residual
       so you can tell if a dip is seasonal (expected) or a real problem.

    2. Customer Cohort Analysis — group customers by first-purchase month,
       track retention over time. Shows whether newer cohorts behave differently.

    3. Anomaly Detection — flag orders with unusual values (amount, qty, or price)
       using the IQR rule (outside Q1 - 1.5*IQR or Q3 + 1.5*IQR).

    4. Market Basket Analysis — find which products are frequently bought together.
       Output: top product pairs by co-occurrence count & lift score.

    5. Sales Forecast — simple exponential smoothing to project next 3 months.
       Good enough for directional planning; not a replacement for Prophet/ARIMA.

    6. Churn Risk Scoring — logistic regression on RFM features to estimate
       probability that each customer will lapse. Uses only pandas + sklearn.

Outputs one Excel file: advanced_analysis_report.xlsx

Design notes:
- No heavy dependencies (no Prophet, no statsmodels, no mlxtend).
- Reuses step3's styling helpers where possible for a consistent look.
- Every analysis degrades gracefully when data is insufficient (returns empty DF).
"""

import os, sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference


# ══════════════════════════════════════════════════════════
# Shared helpers (loader + styling reused from step3 conventions)
# ══════════════════════════════════════════════════════════

def load(folder, fname):
    """Read a CSV if it exists, otherwise return empty DataFrame — lets each
    analysis skip gracefully when a data source isn't available."""
    p = Path(folder) / fname
    if not p.exists():
        return pd.DataFrame()
    return pd.read_csv(p, encoding="utf-8-sig")


# ══════════════════════════════════════════════════════════
# Analysis 1: Seasonality Decomposition
# ══════════════════════════════════════════════════════════

def seasonality_decomposition(monthly: pd.DataFrame) -> pd.DataFrame:
    """Additive decomposition: sales = trend + seasonal + residual.

    - Trend: centered 12-month moving average (smooths out year-over-year bumps)
    - Seasonal: average deviation of each calendar month from the trend,
                then normalized so the 12 values sum to ~0
    - Residual: whatever's left (real noise, one-off events)

    Why additive (not multiplicative)? Simpler, works well when seasonal swings
    are roughly constant in absolute terms. For businesses where seasonality
    grows with scale, multiplicative would be better — but that needs logs.

    Requires at least 13 months of data; returns empty DF if insufficient.
    """
    if monthly.empty or "sales" not in monthly.columns or len(monthly) < 13:
        return pd.DataFrame()

    df = monthly.sort_values("ym").reset_index(drop=True).copy()
    df["sales"] = pd.to_numeric(df["sales"], errors="coerce").fillna(0)

    # Centered 12-month moving average — the "smooth" underlying trend.
    # min_periods=6 allows a partial trend at the edges rather than NaN.
    df["trend"] = df["sales"].rolling(window=12, center=True, min_periods=6).mean()

    # Detrended series = what's left after removing the trend. Contains
    # both the seasonal pattern and random noise.
    df["detrended"] = df["sales"] - df["trend"]

    # Seasonal component: average detrended value for each calendar month,
    # then subtract the overall mean so the 12 seasonal factors sum to ~0.
    df["month"] = df["ym"].astype(str).str[-2:].astype(int)
    month_avg = df.groupby("month")["detrended"].mean()
    month_avg = month_avg - month_avg.mean()
    df["seasonal"] = df["month"].map(month_avg)

    # Residual = actual − trend − seasonal. Big residuals flag real anomalies
    # (promotions, crises, etc.) worth investigating manually.
    df["residual"] = df["sales"] - df["trend"].fillna(0) - df["seasonal"].fillna(0)

    return df[["ym", "sales", "trend", "seasonal", "residual"]]


# ══════════════════════════════════════════════════════════
# Analysis 2: Customer Cohort Analysis
# ══════════════════════════════════════════════════════════

def cohort_analysis(sale_full: pd.DataFrame) -> pd.DataFrame:
    """Retention matrix: rows = cohort (first-purchase month),
                        columns = months since first purchase,
                        cells = # customers still active in that month.

    The first column (month 0) is always the full cohort size.
    Subsequent columns show how many of those customers came back in month 1, 2, …

    Reading the output:
    - Strong diagonal = healthy retention (same customers keep coming back)
    - Rapid drop-off = one-time buyers, need re-engagement
    - Cohorts getting worse over time = product/market fit degrading
    """
    if sale_full.empty or "cusno" not in sale_full.columns or "sdate" not in sale_full.columns:
        return pd.DataFrame()

    df = sale_full[["cusno", "sdate"]].dropna().copy()
    df["sdate"] = pd.to_datetime(df["sdate"], errors="coerce")
    df = df.dropna(subset=["sdate"])
    if df.empty:
        return pd.DataFrame()

    # Each customer's first purchase month defines which cohort they belong to.
    df["order_period"] = df["sdate"].dt.to_period("M")
    df["cohort"] = df.groupby("cusno")["order_period"].transform("min")

    # Cohort index = how many months since the customer's first purchase.
    # Index 0 is the month they joined, 1 is the next month, etc.
    df["cohort_index"] = (df["order_period"] - df["cohort"]).apply(lambda x: x.n)

    # Count unique active customers per (cohort, month-since-start) cell.
    cohort_pivot = (
        df.groupby(["cohort", "cohort_index"])["cusno"]
          .nunique()
          .unstack(fill_value=0)
    )

    # Convert Period index to string for CSV/Excel friendliness.
    cohort_pivot.index = cohort_pivot.index.astype(str)
    cohort_pivot = cohort_pivot.reset_index().rename(columns={"cohort": "cohort_month"})
    return cohort_pivot


# ══════════════════════════════════════════════════════════
# Analysis 3: Anomaly Detection (IQR-based outliers)
# ══════════════════════════════════════════════════════════

def anomaly_detection(sale_full: pd.DataFrame) -> pd.DataFrame:
    """Flag individual sale lines that look unusual on any of these axes:
       - rev (line revenue): promotions, bulk deals, data entry errors
       - prqty (quantity): bulk orders vs accidental double-entries
       - price (unit price): manual overrides, discount abuse

    Uses the classic IQR rule:
       outlier if value < Q1 - 1.5*IQR  or  value > Q3 + 1.5*IQR
    (Q1, Q3 = 25th, 75th percentile; IQR = Q3 - Q1.)

    Why IQR instead of z-score? Retail data is heavily right-skewed (a few
    huge orders). IQR is more robust — doesn't get distorted by the tail.

    Returns only the flagged rows, with a `flag` column listing which
    metric(s) tripped the rule.
    """
    if sale_full.empty:
        return pd.DataFrame()

    df = sale_full.copy()
    numeric_cols = [c for c in ["rev", "prqty", "price"] if c in df.columns]
    if not numeric_cols:
        return pd.DataFrame()

    df["flag"] = ""
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        # Compute bounds on non-null values only, otherwise quantile is NaN
        valid = df[col].dropna()
        if len(valid) < 4:  # need at least a handful of points for quartiles
            continue
        q1, q3 = valid.quantile([0.25, 0.75])
        iqr = q3 - q1
        lo, hi = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        mask = (df[col] < lo) | (df[col] > hi)
        # Append the column name to the flag so users see WHY a row was caught
        df.loc[mask, "flag"] = df.loc[mask, "flag"].where(
            df.loc[mask, "flag"] == "",
            df.loc[mask, "flag"] + ","
        ) + col

    # Keep only anomalies and the columns users actually care about
    flagged = df[df["flag"] != ""].copy()
    keep = [c for c in ["salno", "sdate", "cusno", "cusnm", "prdno", "prdnm",
                        "prqty", "price", "rev", "flag"] if c in flagged.columns]
    return flagged[keep].sort_values("flag")


# ══════════════════════════════════════════════════════════
# Analysis 4: Market Basket Analysis (product co-occurrence + lift)
# ══════════════════════════════════════════════════════════

def market_basket_analysis(sale_full: pd.DataFrame, top_n: int = 30) -> pd.DataFrame:
    """Find which products are frequently bought together in the same order.

    For each unordered pair (A, B):
      - count = number of orders containing BOTH A and B
      - support_A, support_B = how often each product appears on its own
      - lift = P(A AND B) / (P(A) * P(B))
        lift > 1 → bought together more often than chance (cross-sell candidates)
        lift = 1 → independent (no relationship)
        lift < 1 → substitutes (customers pick one or the other)

    Why lift and not just count? A very popular product will pair with
    everything by sheer volume. Lift corrects for that baseline popularity.

    Returns top_n pairs sorted by lift (min count 2 to filter noise).
    """
    required = {"salno", "prdno"}
    if sale_full.empty or not required.issubset(sale_full.columns):
        return pd.DataFrame()

    # One (order, product) row per line; dedupe in case the same product
    # appears twice on one order (it shouldn't affect pair counts).
    lines = sale_full[["salno", "prdno"]].dropna().drop_duplicates()
    if len(lines) < 2:
        return pd.DataFrame()

    # Optional product name lookup for readable output
    name_map = {}
    if "prdnm" in sale_full.columns:
        name_map = (
            sale_full.dropna(subset=["prdno"])
                    .drop_duplicates("prdno")
                    .set_index("prdno")["prdnm"]
                    .to_dict()
        )

    total_orders = lines["salno"].nunique()
    if total_orders == 0:
        return pd.DataFrame()

    # Support per product: fraction of orders containing this product
    prod_support = (
        lines.groupby("prdno")["salno"].nunique() / total_orders
    ).to_dict()

    # Build pair counts via self-join on salno, then keep each unordered pair once
    # (A,B where A < B alphabetically — eliminates (A,A) and (B,A) duplicates).
    joined = lines.merge(lines, on="salno", suffixes=("_a", "_b"))
    pairs = joined[joined["prdno_a"] < joined["prdno_b"]]
    if pairs.empty:
        return pd.DataFrame()

    pair_counts = (
        pairs.groupby(["prdno_a", "prdno_b"])["salno"]
             .nunique()
             .reset_index(name="count")
    )
    # Filter out pairs that only co-occur once — too noisy to be meaningful.
    pair_counts = pair_counts[pair_counts["count"] >= 2]
    if pair_counts.empty:
        return pd.DataFrame()

    # Compute lift for each pair
    pair_counts["support_pair"] = pair_counts["count"] / total_orders
    pair_counts["support_a"] = pair_counts["prdno_a"].map(prod_support)
    pair_counts["support_b"] = pair_counts["prdno_b"].map(prod_support)
    pair_counts["lift"] = (
        pair_counts["support_pair"]
        / (pair_counts["support_a"] * pair_counts["support_b"])
    )
    # Attach product names if available
    pair_counts["name_a"] = pair_counts["prdno_a"].map(name_map).fillna("")
    pair_counts["name_b"] = pair_counts["prdno_b"].map(name_map).fillna("")

    return (
        pair_counts.sort_values("lift", ascending=False)
                  .head(top_n)
                  [["prdno_a", "name_a", "prdno_b", "name_b", "count", "lift"]]
                  .reset_index(drop=True)
    )


# ══════════════════════════════════════════════════════════
# Analysis 5: Sales Forecast (exponential smoothing + seasonal adjustment)
# ══════════════════════════════════════════════════════════

def sales_forecast(monthly: pd.DataFrame, horizon: int = 3) -> pd.DataFrame:
    """Project the next `horizon` months of sales.

    Method selection based on available history:
      - >= 24 months: trend (EWMA) + seasonal adjustment (month-of-year avg deviation)
      - >=  6 months: trend only (EWMA extrapolation)
      -  <  6 months: not enough data → return empty

    Why EWMA? Recent months weigh more than old ones (reacts to regime shifts).
    Why not ARIMA/Prophet? Those need extra deps and tuning. For a 3-month
    directional forecast, EWMA + seasonal adjust is usually within 10-15% of
    more complex models and far easier to audit.

    Returns a DataFrame with columns: ym, sales (actual or forecast), type.
    type = "actual" for historical rows, "forecast" for projected rows.
    """
    if monthly.empty or "sales" not in monthly.columns or len(monthly) < 6:
        return pd.DataFrame()

    df = monthly.sort_values("ym").reset_index(drop=True).copy()
    df["sales"] = pd.to_numeric(df["sales"], errors="coerce").fillna(0)
    df["type"] = "實際"

    # EWMA trend: alpha=0.3 → recent month = 30% weight, long tail = 70%.
    # This is a reasonable middle ground; increase for more reactive forecasts.
    trend = df["sales"].ewm(alpha=0.3, adjust=False).mean()
    last_trend = float(trend.iloc[-1])
    # Use last-6-month trend slope for gentle extrapolation (not pure flat line).
    recent_slope = float((trend.iloc[-1] - trend.iloc[-6]) / 6) if len(trend) >= 6 else 0.0

    # Seasonal factors (only if we have >= 2 full years)
    seasonal_map = {}
    if len(df) >= 24:
        df["month"] = df["ym"].astype(str).str[-2:].astype(int)
        # Deviation of each month from the trend, averaged across years
        df["detrended"] = df["sales"] - trend
        month_avg = df.groupby("month")["detrended"].mean()
        month_avg = month_avg - month_avg.mean()  # center on zero
        seasonal_map = month_avg.to_dict()

    # Parse last ym to roll forward the period
    last_ym = str(df["ym"].iloc[-1])
    try:
        last_period = pd.Period(last_ym, freq="M")
    except Exception:
        return df[["ym", "sales", "type"]]

    # Build forecast rows
    forecast_rows = []
    for step in range(1, horizon + 1):
        next_period = last_period + step
        month_num = next_period.month
        projected = last_trend + recent_slope * step
        if seasonal_map:
            projected += seasonal_map.get(month_num, 0.0)
        forecast_rows.append({
            "ym": str(next_period),
            "sales": max(0.0, projected),  # sales can't go negative
            "type": "預測",
        })

    out = pd.concat(
        [df[["ym", "sales", "type"]], pd.DataFrame(forecast_rows)],
        ignore_index=True,
    )
    return out


# ══════════════════════════════════════════════════════════
# Analysis 6: Churn Risk Scoring (logistic regression on RFM features)
# ══════════════════════════════════════════════════════════

def churn_risk_scoring(cust_agg: pd.DataFrame) -> pd.DataFrame:
    """Estimate probability each customer will churn.

    Labeling approach (no explicit churn column in the data):
      A customer is labeled "churned" if their R_days (days since last
      transaction) falls in the top 25% of the distribution. This is a
      heuristic — the real definition depends on the business's sales cycle.

    Features: R_days, order_count, sales, avg_order_value (if available).
    Model: logistic regression (sklearn), which is interpretable and stable
           on small datasets. The `churn_prob` column is the predicted
           probability — customers with prob > 0.5 are high risk.

    Returns cust_agg enriched with `churn_prob` and `churn_flag` columns,
    sorted by risk descending. Returns empty DataFrame if sklearn missing
    or data too small (< 20 customers for any reliable split).
    """
    if cust_agg.empty or "last_transaction" not in cust_agg.columns:
        return pd.DataFrame()
    if len(cust_agg) < 20:
        # Logistic regression needs enough variety; small-sample scores would
        # be meaningless. Below this threshold, skip gracefully.
        return pd.DataFrame()

    try:
        from sklearn.linear_model import LogisticRegression
        from sklearn.preprocessing import StandardScaler
    except ImportError:
        # sklearn is in requirements.txt, but fail gracefully in minimal installs
        return pd.DataFrame()

    df = cust_agg.copy()
    df["last_transaction"] = pd.to_datetime(df["last_transaction"], errors="coerce")
    df = df.dropna(subset=["last_transaction"])
    if df.empty:
        return pd.DataFrame()

    # R_days relative to the latest transaction in the dataset
    ref_date = df["last_transaction"].max()
    df["R_days"] = (ref_date - df["last_transaction"]).dt.days

    # Label: top-quartile R_days = churned (no contact for a long time)
    threshold = df["R_days"].quantile(0.75)
    df["is_churned"] = (df["R_days"] >= threshold).astype(int)
    # If the label is degenerate (all same class), can't train — bail out
    if df["is_churned"].nunique() < 2:
        return pd.DataFrame()

    feature_cols = [c for c in ["R_days", "order_count", "sales", "avg_order_value"]
                    if c in df.columns]
    if len(feature_cols) < 2:
        return pd.DataFrame()

    # Convert features to numeric, fill NaN with 0 (missing = low activity)
    X = df[feature_cols].apply(pd.to_numeric, errors="coerce").fillna(0).values
    y = df["is_churned"].values

    # Standardize so all features contribute fairly regardless of scale
    # (R_days in 100s, order_count in 10s, sales in 1000s without this)
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # max_iter bumped up so convergence warning doesn't spam on small data
    model = LogisticRegression(max_iter=1000, random_state=42)
    model.fit(X_scaled, y)

    df["churn_prob"] = model.predict_proba(X_scaled)[:, 1]
    df["churn_flag"] = np.where(df["churn_prob"] >= 0.5, "高風險", "低風險")

    keep = [c for c in ["cusno", "cusnm", "order_count", "sales",
                        "R_days", "churn_prob", "churn_flag"] if c in df.columns]
    return df[keep].sort_values("churn_prob", ascending=False).reset_index(drop=True)


# ══════════════════════════════════════════════════════════
# Excel sheet writers — reuse step3's styling helpers for visual consistency
# ══════════════════════════════════════════════════════════
# Lazy import inside functions to avoid hard dependency at import time
# (so the analysis functions alone can be reused without openpyxl styling).

def _import_step3():
    """Dynamic import so step5 can run even if step3 is moved/renamed later."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "step3_analyze", Path(__file__).parent / "step3_analyze.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def sheet_seasonality(wb, monthly, s3):
    """Draw decomposition table + stacked line chart (actual/trend/seasonal)."""
    ws = wb.create_sheet("季節性分析")
    s3.page_title(ws, "季節性分解", 5)

    df = seasonality_decomposition(monthly)
    if df.empty:
        ws["A2"] = "需至少13個月資料才能進行分解"; return

    for ci, h in enumerate(["月份", "銷售額", "趨勢", "季節性", "殘差"], 1):
        s3.hdr(ws, 3, ci, h)
    for i, r in df.reset_index(drop=True).iterrows():
        row = i + 4
        bg = s3.GRAY if i % 2 == 0 else s3.WHITE
        s3.cel(ws, row, 1, r["ym"], align="center", bg=bg)
        s3.cel(ws, row, 2, round(float(r["sales"]), 0),    "#,##0", bg=bg)
        s3.cel(ws, row, 3, round(float(r["trend"] or 0), 0),   "#,##0", bg=bg)
        s3.cel(ws, row, 4, round(float(r["seasonal"] or 0), 0), "#,##0", bg=bg)
        s3.cel(ws, row, 5, round(float(r["residual"] or 0), 0), "#,##0", bg=bg)
    s3.border(ws, 3, 3 + len(df), 1, 5)

    # Line chart: actual vs trend (on-sheet, anchored right of the table)
    n = len(df)
    chart = LineChart()
    chart.title = "銷售額 vs 趨勢"
    chart.height = 10; chart.width = 20
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=3 + n)
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    chart.set_categories(cats)
    ws.add_chart(chart, "G3")
    s3.widths(ws, {"A": 12, "B": 14, "C": 14, "D": 14, "E": 14})


def sheet_cohort(wb, sale_full, s3):
    """Retention matrix: cohort month (rows) × months-since-join (cols)."""
    ws = wb.create_sheet("同期群分析")
    s3.page_title(ws, "客戶同期群留存率", 12)

    df = cohort_analysis(sale_full)
    if df.empty:
        ws["A2"] = "無同期群資料（需 sale_full.csv 含 cusno + sdate）"; return

    cohort_cols = [c for c in df.columns if c != "cohort_month"][:13]
    headers = ["加入月份"] + [f"第{c}月" for c in cohort_cols]
    for ci, h in enumerate(headers, 1):
        s3.hdr(ws, 3, ci, h)

    # Heatmap-style shading: darker = more retained customers (relative to M0)
    for i, r in df.iterrows():
        row = i + 4
        m0 = float(r[cohort_cols[0]]) if cohort_cols else 0.0
        s3.cel(ws, row, 1, str(r["cohort_month"]), align="center")
        for ci, c in enumerate(cohort_cols, 2):
            val = float(r[c]) if pd.notna(r[c]) else 0
            # Retention ratio drives background intensity; M0 always = 1.0 (full cohort)
            ratio = val / m0 if m0 > 0 else 0
            # Simple 3-tier coloring — green for high retention, yellow mid, red low
            if ratio >= 0.5:   bg = "C6EFCE"
            elif ratio >= 0.2: bg = "FFEB9C"
            elif ratio > 0:    bg = "FFC7CE"
            else:              bg = s3.WHITE
            s3.cel(ws, row, ci, int(val), "#,##0", bg=bg, align="center")
    s3.border(ws, 3, 3 + len(df), 1, len(headers))
    s3.widths(ws, {get_column_letter(i): 10 for i in range(1, len(headers) + 1)})


def sheet_anomaly(wb, sale_full, s3):
    """List of IQR-flagged sale lines, sorted by which metric tripped the rule."""
    ws = wb.create_sheet("異常偵測")
    s3.page_title(ws, "異常銷售明細（IQR離群值）", 10)

    df = anomaly_detection(sale_full)
    if df.empty:
        ws["A2"] = "未偵測到異常（或資料不足）"; return

    cols = list(df.columns)
    hdrs_map = {"salno": "銷售單號", "sdate": "日期", "cusno": "客戶編號",
                "cusnm": "客戶名稱", "prdno": "產品編號", "prdnm": "產品名稱",
                "prqty": "數量", "price": "單價", "rev": "金額", "flag": "異常欄位"}
    for ci, c in enumerate(cols, 1):
        s3.hdr(ws, 3, ci, hdrs_map.get(c, c))
    for i, r in df.reset_index(drop=True).iterrows():
        row = i + 4
        # Every flagged row gets a pink background — it's already an exception
        for ci, c in enumerate(cols, 1):
            v = r.get(c, "")
            fmt = "#,##0" if c in ["prqty", "price", "rev"] else None
            s3.cel(ws, row, ci, v, fmt=fmt, bg="FFE4E1",
                   align="left" if c in ["cusnm", "prdnm", "flag", "salno", "cusno", "prdno"] else "right")
    s3.border(ws, 3, 3 + len(df), 1, len(cols))
    s3.widths(ws, {get_column_letter(i): 14 for i in range(1, len(cols) + 1)})


def sheet_basket(wb, sale_full, s3):
    """Top product pairs by lift — the cross-sell candidate list."""
    ws = wb.create_sheet("購物籃分析")
    s3.page_title(ws, "購物籃分析（依關聯度排序）", 6)

    df = market_basket_analysis(sale_full)
    if df.empty:
        ws["A2"] = "重複搭配不足，無法計算關聯度"; return

    for ci, h in enumerate(["排名", "產品A", "產品A名稱", "產品B", "產品B名稱", "共同出現次數", "關聯度"], 1):
        s3.hdr(ws, 3, ci, h)
    for i, r in df.iterrows():
        row = i + 4
        bg = s3.GRAY if i % 2 == 0 else s3.WHITE
        # Highlight strong associations (lift > 1.5) in green — these are the
        # most actionable cross-sell recommendations for marketing/bundle teams
        lift_val = float(r["lift"])
        lift_bg = "C6EFCE" if lift_val > 1.5 else bg
        s3.cel(ws, row, 1, i + 1,             align="center", bg=bg)
        s3.cel(ws, row, 2, r["prdno_a"],      align="center", bg=bg)
        s3.cel(ws, row, 3, r["name_a"],       align="left",   bg=bg)
        s3.cel(ws, row, 4, r["prdno_b"],      align="center", bg=bg)
        s3.cel(ws, row, 5, r["name_b"],       align="left",   bg=bg)
        s3.cel(ws, row, 6, int(r["count"]),   "#,##0",        bg=bg)
        s3.cel(ws, row, 7, round(lift_val, 2), "0.00",         bg=lift_bg)
    s3.border(ws, 3, 3 + len(df), 1, 7)
    s3.widths(ws, {"A": 6, "B": 12, "C": 24, "D": 12, "E": 24, "F": 10, "G": 10})


def sheet_forecast(wb, monthly, s3):
    """Historical sales + projected next-N months with clear visual separation."""
    ws = wb.create_sheet("銷售預測")
    s3.page_title(ws, "銷售預測（指數平滑+季節調整）", 4)

    df = sales_forecast(monthly, horizon=3)
    if df.empty:
        ws["A2"] = "需至少6個月資料才能預測"; return

    for ci, h in enumerate(["月份", "銷售額", "類型"], 1):
        s3.hdr(ws, 3, ci, h)
    for i, r in df.reset_index(drop=True).iterrows():
        row = i + 4
        is_forecast = r["type"] == "預測"
        bg = "FFE4B5" if is_forecast else (s3.GRAY if i % 2 == 0 else s3.WHITE)
        s3.cel(ws, row, 1, r["ym"],                    align="center", bg=bg)
        s3.cel(ws, row, 2, round(float(r["sales"]), 0), "#,##0",        bg=bg)
        s3.cel(ws, row, 3, r["type"],                   align="center", bg=bg)
    s3.border(ws, 3, 3 + len(df), 1, 3)

    # Line chart of full series — the eye catches the forecast tail immediately
    n = len(df)
    chart = LineChart()
    chart.title = "歷史銷售 + 預測"
    chart.height = 10; chart.width = 20
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=3 + n)
    chart.add_data(data_ref, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    chart.set_categories(cats)
    ws.add_chart(chart, "F3")
    s3.widths(ws, {"A": 12, "B": 14, "C": 12})


def sheet_churn(wb, cust_agg, s3):
    """Customer list sorted by churn probability with risk flag."""
    ws = wb.create_sheet("流失風險")
    s3.page_title(ws, "客戶流失風險評分", 7)

    df = churn_risk_scoring(cust_agg)
    if df.empty:
        ws["A2"] = "需>=20位客戶且距上次購買天數有差異才能訓練流失模型"; return

    cols = list(df.columns)
    hdrs_map = {"cusno": "客戶編號", "cusnm": "客戶名稱", "order_count": "訂單數",
                "sales": "銷售額", "R_days": "距上次購買天數",
                "churn_prob": "流失機率", "churn_flag": "風險等級"}
    for ci, c in enumerate(cols, 1):
        s3.hdr(ws, 3, ci, hdrs_map.get(c, c))
    for i, r in df.iterrows():
        row = i + 4
        # High risk rows get red-ish background; low risk stays neutral.
        # Makes the top of the sheet unmistakably actionable.
        is_high = r.get("churn_flag", "") == "高風險"
        bg = "FFC7CE" if is_high else (s3.GRAY if i % 2 == 0 else s3.WHITE)
        for ci, c in enumerate(cols, 1):
            v = r.get(c, "")
            if c == "churn_prob":
                s3.cel(ws, row, ci, round(float(v), 3), "0.0%", bg=bg)
            elif c in ["order_count", "sales", "R_days"]:
                s3.cel(ws, row, ci, v, "#,##0", bg=bg)
            else:
                align = "left" if c in ["cusnm", "churn_flag"] else "center"
                s3.cel(ws, row, ci, v, align=align, bg=bg)
    s3.border(ws, 3, 3 + len(df), 1, len(cols))
    s3.widths(ws, {"A": 12, "B": 24, "C": 10, "D": 14, "E": 14, "F": 12, "G": 12})


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = input("請輸入彙總資料夾路徑 (step2 輸出):\n> ").strip()

    src = Path(folder)
    if not src.exists():
        print(f"找不到資料夾: {folder}")
        input("按 Enter 離開..."); return

    out_path = src.parent / "advanced_analysis_report.xlsx"
    print(f"\nSource: {src}")
    print(f"Output: {out_path}\n{'='*50}")

    # Load inputs: some analyses use the aggregate tables, others need the
    # line-level sale_full (which has salno + prdno + prdnm for basket analysis).
    monthly   = load(src, "monthly.csv")
    cust_agg  = load(src, "cust_agg.csv")
    sale_full = load(src, "sale_full.csv")

    print("載入報表樣式...")
    try:
        s3 = _import_step3()
    except Exception as e:
        print(f"無法載入 step3: {e}")
        input("按 Enter 離開..."); return

    print("建立進階報表中...")
    wb = Workbook()
    wb.remove(wb.active)

    # Order chosen so the executive-friendly sheets come first,
    # deep-dive tables last.
    sheet_forecast(wb, monthly, s3)       # What's next?
    sheet_seasonality(wb, monthly, s3)     # Why did this month dip?
    sheet_cohort(wb, sale_full, s3)        # Are we keeping customers?
    sheet_churn(wb, cust_agg, s3)          # Who's about to leave?
    sheet_basket(wb, sale_full, s3)        # What should we bundle?
    sheet_anomaly(wb, sale_full, s3)       # What looks wrong?

    wb.save(out_path)
    print(f"\n進階報表完成: {out_path}")
    print("\n工作表:")
    for ws in wb.worksheets:
        print(f"   {ws.title}")
    input("\n按 Enter 離開...")


if __name__ == "__main__":
    main()
