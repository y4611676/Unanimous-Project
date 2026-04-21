"""
Step 4: De-identified Report
- Read aggregated/ data
- Apply consistent masking to identifier columns
- Apply proportional scaling to monetary values (preserves trends/ratios but not actual scale)
- Apply consistent date shifting (preserves relative timeline but not actual dates)
- Output de-identified Excel report (does not modify source data)

Masking rules:
  cusno / cusnm  →  CUST-001 / Customer A
  prdno / prdnm  →  PROD-001 / Product A
  facno / facnm  →  SUPP-001 / Supplier A
  monetary cols   →  multiply by random scale factor (fixed per run, controlled by ANON_SEED)
  date cols       →  shift by random months (same ANON_SEED)
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from dateutil.relativedelta import relativedelta
import importlib.util

# ══════════════════════════════════════════════════════════
# De-identification parameters (change ANON_SEED for different scale/shift)
# ══════════════════════════════════════════════════════════
ANON_SEED = 20240101

_rng = np.random.default_rng(ANON_SEED)

# Money scale factor: two branches to guarantee >=40% deviation from real values.
# If too close to 1.0, someone could guess the original amounts from the report.
# Branch 0: shrink to 15-55% of real → values look like a smaller company
# Branch 1: expand to 150-280% of real → values look like a bigger company
_side = int(_rng.integers(0, 2))
MONEY_SCALE = round(
    float(_rng.uniform(0.15, 0.55)) if _side == 0
    else float(_rng.uniform(1.5, 2.8)),
    6
)

# Date shift: 24-72 months (2-6 years) in a random direction.
# Must be large enough that you can't map back to real calendar events.
_dir = 1 if int(_rng.integers(0, 2)) == 0 else -1
DATE_SHIFT_M = _dir * int(_rng.integers(24, 73))

# ── Monetary columns to scale ─────────────────────────────
MONEY_COLS = {
    "monthly":   ["sales","purchases","gross_profit","balance","sales_detail","cost"],
    "cust_agg":  ["sales","avg_order_value"],
    "prod_agg":  ["sales","cost","gross_profit"],
    "fact_agg":  ["purchases","avg_purchase_price"],
    "stock_agg": ["stock_value","price","pcost"],
}

# ── Date columns to shift ────────────────────────────────
DATE_COLS = {
    "cust_agg":  ["first_transaction","last_transaction"],
    "fact_agg":  ["first_purchase","last_purchase"],
}


# ══════════════════════════════════════════════════════════
# Dynamically import step3 report functions
# ══════════════════════════════════════════════════════════
def _import_step3():
    spec = importlib.util.spec_from_file_location(
        "step3_analyze",
        Path(__file__).parent / "step3_analyze.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ══════════════════════════════════════════════════════════
# Masking utilities
# ══════════════════════════════════════════════════════════

def _idx_to_letters(i):
    return chr(65 + i) if i < 26 else chr(65 + i // 26 - 1) + chr(65 + i % 26)


def _make_id_map(series, prefix):
    """Original ID → XXXX-001 mapping (string keys)"""
    vals = [str(v).strip() for v in series.dropna().unique() if str(v).strip() not in ("", "nan")]
    return {v: f"{prefix}-{str(i+1).zfill(3)}" for i, v in enumerate(vals)}


def _make_nm_map(id_series, nm_series, label):
    """Build name→masked_name mapping based on id column order.
    Names sharing the same id get the same masked label (e.g. 'Customer A').
    This ensures consistency: if cusno C001='Acme Corp' appears in multiple tables,
    it always maps to the same masked name."""
    vals = [str(v).strip() for v in id_series.dropna().unique() if str(v).strip() not in ("", "nan")]
    nm_map = {}
    for i, orig_id in enumerate(vals):
        mask = id_series.astype(str).str.strip() == orig_id
        names = nm_series[mask].dropna().unique()
        for nm in names:
            nm_str = str(nm).strip()
            if nm_str and nm_str != "nan":
                nm_map[nm_str] = f"{label}{_idx_to_letters(i)}"
    return nm_map


def _apply(df, col_maps):
    """Apply col_maps (column→mapping dict) to a df copy, coercing to string for comparison."""
    df = df.copy()
    for col, mapping in col_maps.items():
        if col not in df.columns:
            continue
        df[col] = df[col].apply(
            lambda x: mapping.get(str(x).strip(), x)
            if pd.notna(x) and str(x).strip() not in ("", "nan")
            else x
        )
    return df


def build_id_maps(cust_agg, prod_agg, fact_agg):
    """Build masking maps for all identifier columns, return col_maps dict."""
    col_maps = {}

    if not cust_agg.empty and "cusno" in cust_agg.columns:
        col_maps["cusno"] = _make_id_map(cust_agg["cusno"], "CUST")
        if "cusnm" in cust_agg.columns:
            col_maps["cusnm"] = _make_nm_map(cust_agg["cusno"], cust_agg["cusnm"], "客戶 ")

    if not prod_agg.empty and "prdno" in prod_agg.columns:
        col_maps["prdno"] = _make_id_map(prod_agg["prdno"], "PROD")
        if "prdnm" in prod_agg.columns:
            col_maps["prdnm"] = _make_nm_map(prod_agg["prdno"], prod_agg["prdnm"], "商品 ")

    if not fact_agg.empty and "facno" in fact_agg.columns:
        col_maps["facno"] = _make_id_map(fact_agg["facno"], "SUPP")
        if "facnm" in fact_agg.columns:
            col_maps["facnm"] = _make_nm_map(fact_agg["facno"], fact_agg["facnm"], "供應商 ")

    return col_maps


# ══════════════════════════════════════════════════════════
# Money scaling
# ══════════════════════════════════════════════════════════

def scale_money(df, df_key):
    # Proportional scaling: multiply all monetary columns by the same factor.
    # This preserves ratios (e.g. GP rate stays the same) but hides true magnitude.
    df = df.copy()
    for col in MONEY_COLS.get(df_key, []):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0) * MONEY_SCALE
    return df


# ══════════════════════════════════════════════════════════
# Date shifting
# ══════════════════════════════════════════════════════════

def _shift_ym_str(ym_str):
    """
    Shift "2019-06" or "2019Q2" style strings by DATE_SHIFT_M months.
    Returns original value if unparseable.
    """
    s = str(ym_str).strip()
    try:
        # Handle 2019-06 format
        if len(s) == 7 and s[4] == "-":
            dt = pd.to_datetime(s + "-01") + relativedelta(months=DATE_SHIFT_M)
            return dt.strftime("%Y-%m")
        # Handle 2019Q2 format
        if len(s) == 6 and s[4] == "Q":
            base_month = (int(s[5]) - 1) * 3 + 1
            dt = pd.to_datetime(f"{s[:4]}-{base_month:02d}-01") + relativedelta(months=DATE_SHIFT_M)
            new_q = (dt.month - 1) // 3 + 1
            return f"{dt.year}Q{new_q}"
    except Exception:
        pass
    return ym_str


def _shift_date_col(series):
    """Shift a date column (string or datetime)."""
    result = pd.to_datetime(series, errors="coerce")
    shifted = result.apply(
        lambda d: d + relativedelta(months=DATE_SHIFT_M) if pd.notna(d) else d
    )
    return shifted.dt.strftime("%Y-%m-%d").where(shifted.notna(), other="")


def shift_dates(df, df_key):
    # Consistent shift: all dates move by the same offset, so relative timeline
    # (seasonality, gaps, sequences) is perfectly preserved — only the absolute
    # calendar position changes, preventing re-identification by date matching.
    df = df.copy()
    for col in DATE_COLS.get(df_key, []):
        if col in df.columns:
            df[col] = _shift_date_col(df[col])
    # Special handling for monthly ym / yq columns
    if df_key == "monthly":
        for col in ["ym", "yq"]:
            if col in df.columns:
                df[col] = df[col].apply(_shift_ym_str)
    return df


# ══════════════════════════════════════════════════════════
# Mapping sheet
# ══════════════════════════════════════════════════════════

def _add_mapping_sheet(wb, col_maps, cust_agg, prod_agg, fact_agg):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet("對照表（機密）")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = (
        "匿名對照（內部使用） | "
        f"金額縮放：{MONEY_SCALE:.4f}  "
        f"日期位移：{DATE_SHIFT_M:+d} 個月  "
        f"Seed：{ANON_SEED}"
    )
    c.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="C00000")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    row = 3
    sections = [
        ("客戶對照", "cusno", "cusnm", cust_agg),
        ("商品對照", "prdno", "prdnm", prod_agg),
        ("供應商對照", "facno", "facnm", fact_agg),
    ]

    for section_title, id_key, nm_key, ref_df in sections:
        if id_key not in col_maps:
            continue
        ws.cell(row=row, column=1, value=section_title).font = Font(bold=True, size=11)
        row += 1
        for ci, h in enumerate(["原始編號","匿名編號","原始名稱","匿名名稱"], 1):
            ws.cell(row=row, column=ci, value=h).font = Font(bold=True)
        row += 1

        id_map = col_maps[id_key]
        nm_map = col_maps.get(nm_key, {})
        nm_rev = {v: k for k, v in nm_map.items()}

        for orig_id, anon_id in id_map.items():
            idx = list(id_map.keys()).index(orig_id)
            anon_nm = f"{'客戶 ' if id_key=='cusno' else '商品 ' if id_key=='prdno' else '供應商 '}{_idx_to_letters(idx)}"
            orig_nm = nm_rev.get(anon_nm, "")
            ws.cell(row=row, column=1, value=orig_id)
            ws.cell(row=row, column=2, value=anon_id)
            ws.cell(row=row, column=3, value=orig_nm)
            ws.cell(row=row, column=4, value=anon_nm if orig_nm else "")
            row += 1
        row += 2

    for col_letter, w in [("A",18),("B",14),("C",30),("D",14)]:
        ws.column_dimensions[col_letter].width = w


# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = input("Enter path to aggregated folder (step2 output):\n> ").strip()

    src = Path(folder)
    if not src.exists():
        print(f"Folder not found: {folder}")
        input("Press Enter to exit..."); return

    out_path = src.parent / "business_analysis_deidentified.xlsx"
    print(f"\nSource: {src}")
    print(f"Output: {out_path}\n{'='*50}")
    print(f"Seed={ANON_SEED}  Money Scale={MONEY_SCALE:.4f}  Date Shift={DATE_SHIFT_M:+d} months")

    # ── Load step3 report module ─────────────────────────
    print("\nLoading report module (step3_analyze.py)...")
    try:
        s3 = _import_step3()
    except Exception as e:
        print(f"Failed to load step3_analyze.py: {e}")
        input("Press Enter to exit..."); return

    # ── Read aggregated data ─────────────────────────────
    def load(fname):
        p = src / fname
        if not p.exists(): return pd.DataFrame()
        return pd.read_csv(p, encoding="utf-8-sig")

    monthly   = load("monthly.csv")
    cust_agg  = load("cust_agg.csv")
    prod_agg  = load("prod_agg.csv")
    fact_agg  = load("fact_agg.csv")
    stock_agg = load("stock_agg.csv")

    # Numeric column conversion
    for df, cols in [
        (monthly,  ["sales","purchases","gross_profit","gp_rate","balance","sales_detail"]),
        (cust_agg, ["sales","order_count","avg_order_value","transaction_months"]),
        (prod_agg, ["sales","cost","gross_profit","gp_rate","sales_qty"]),
        (stock_agg,["qty","safeqty","stock_value"]),
    ]:
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # ── Build ID masks ───────────────────────────────────
    print("\nBuilding ID masks...")
    col_maps = build_id_maps(cust_agg, prod_agg, fact_agg)
    print(f"   Customers: {len(col_maps.get('cusno', {}))} entries")
    print(f"   Products: {len(col_maps.get('prdno', {}))} entries")
    print(f"   Suppliers: {len(col_maps.get('facno', {}))} entries")

    # ── Apply masks ──────────────────────────────────────
    print("Applying ID masks...")
    anon_cust  = _apply(cust_agg,  col_maps)
    anon_prod  = _apply(prod_agg,  col_maps)
    anon_fact  = _apply(fact_agg,  col_maps)
    anon_stock = _apply(stock_agg, col_maps)

    # ── Scale monetary columns ───────────────────────────
    print("Scaling monetary columns...")
    anon_monthly  = scale_money(monthly,   "monthly")
    anon_cust     = scale_money(anon_cust, "cust_agg")
    anon_prod     = scale_money(anon_prod, "prod_agg")
    anon_fact     = scale_money(anon_fact, "fact_agg")
    anon_stock    = scale_money(anon_stock,"stock_agg")

    # ── Shift date columns ───────────────────────────────
    print("Shifting date columns...")
    anon_monthly = shift_dates(anon_monthly, "monthly")
    anon_cust    = shift_dates(anon_cust,    "cust_agg")
    anon_fact    = shift_dates(anon_fact,    "fact_agg")

    # ── RFM analysis ─────────────────────────────────────
    print("RFM analysis...")
    rfm_df = s3.rfm_analysis(anon_cust)

    # ── Generate report ──────────────────────────────────
    print("Building report...")
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    s3.sheet_charts(wb, anon_monthly, anon_cust, anon_prod)
    s3.sheet_summary(wb, anon_monthly, anon_cust, anon_prod, anon_stock)
    s3.sheet_rfm(wb, rfm_df)
    s3.sheet_pareto(wb, anon_prod, anon_cust)
    s3.sheet_gross_margin(wb, anon_prod)
    s3.sheet_inventory_alert(wb, anon_stock)
    _add_mapping_sheet(wb, col_maps, cust_agg, prod_agg, fact_agg)

    wb.save(out_path)
    print(f"\nDe-identified report complete: {out_path}")
    print("\nSheets:")
    for ws in wb.worksheets:
        print(f"   {ws.title}")
    print("\nBefore sharing externally, delete the '對照表（機密）' sheet!")
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
