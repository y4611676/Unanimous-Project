"""
Step 3: Data Analysis → Excel Report
- Read aggregated/ data from step2
- Perform RFM, Pareto, trend, inventory alert analyses
- Output full Excel report
"""

import os, sys
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ── Colors ────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
GREEN      = "375623"
GREEN_SOFT = "70AD47"
RED        = "C00000"
ORANGE     = "ED7D31"
GRAY       = "F2F2F2"
WHITE      = "FFFFFF"
BLACK      = "000000"

# ── Style utilities ───────────────────────────────────────

def hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, sz=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return c

def cel(ws, row, col, val, fmt=None, bold=False, fg=BLACK, bg=None, align="right"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=fg, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: c.number_format = fmt
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    return c

def border(ws, r1, r2, c1, c2):
    thin = Side(style="thin", color="BFBFBF")
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def title(ws, row, col, text, span, bg=MED_BLUE, sz=11):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22

def page_title(ws, text, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws["A1"]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.sheet_view.showGridLines = False

def load(folder, fname):
    p = Path(folder) / fname
    if not p.exists(): return pd.DataFrame()
    return pd.read_csv(p, encoding="utf-8-sig")

def widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

# ══════════════════════════════════════════════════════════
# Analysis functions
# ══════════════════════════════════════════════════════════

def rfm_analysis(cust_agg):
    """RFM customer segmentation — scores each customer on three axes:
       R (Recency): days since last purchase — lower is better, so scoring is reversed
       F (Frequency): total order count — higher is better
       M (Monetary): total spend — higher is better
    Each axis is split into quartiles (1-4) via pd.qcut; combined into a 3-digit score."""
    if cust_agg.empty: return pd.DataFrame()
    df = cust_agg.copy()
    df["last_transaction"] = pd.to_datetime(df["last_transaction"], errors="coerce")
    ref_date = df["last_transaction"].max()
    df["R_days"] = (ref_date - df["last_transaction"]).dt.days
    df["F"] = df["order_count"]
    df["M"] = df["sales"]

    # qcut splits into 4 equal-sized bins; R is reversed (fewer days = better score)
    for col, label in [("R_days","R_score"),("F","F_score"),("M","M_score")]:
        try:
            df[label] = pd.qcut(df[col], q=4, labels=[4,3,2,1] if col=="R_days" else [1,2,3,4], duplicates="drop")
        except:
            df[label] = 2
    df["RFM_score"] = df["R_score"].astype(str) + df["F_score"].astype(str) + df["M_score"].astype(str)

    # Segment mapping — business meaning:
    #   High Value: recent + frequent + big spender → protect at all costs
    #   Loyal: recent + frequent → nurture, upsell opportunities
    #   Dormant: not recent but was frequent → win-back campaigns
    #   Big Spender: recent + high spend but infrequent → encourage repeat
    #   At Risk: not recent + infrequent → about to churn, act fast
    #   Regular: everyone else → standard engagement
    def segment(row):
        r,f,m = int(str(row["R_score"])), int(str(row["F_score"])), int(str(row["M_score"]))
        if r >= 3 and f >= 3 and m >= 3: return "高價值"
        if r >= 3 and f >= 3:            return "忠誠客戶"
        if r <= 2 and f >= 3:            return "沉睡客戶"
        if r >= 3 and m >= 3:            return "大客戶"
        if r <= 2 and f <= 2:            return "流失風險"
        return "一般客戶"

    df["segment"] = df.apply(segment, axis=1)
    return df

def pareto_analysis(df, value_col, name_col):
    """Pareto 80/20 analysis — ranks items by value_col descending, computes
    cumulative share, and tags each as 'Top 80%' or 'Bottom 20%'.
    In most businesses, ~20% of customers/products drive ~80% of revenue."""
    df = df.sort_values(value_col, ascending=False).copy()
    total = df[value_col].sum()
    df["share"] = df[value_col] / total
    df["cumulative_share"] = df["share"].cumsum()
    df["pareto"] = df["cumulative_share"].apply(lambda x: "前80%" if x <= 0.8 else "後20%")
    return df

# ══════════════════════════════════════════════════════════
# Report sheets
# ══════════════════════════════════════════════════════════

def sheet_summary(wb, monthly, cust_agg, prod_agg, stock_agg):
    # Executive dashboard: 6 KPI cards + monthly trend table
    ws = wb.create_sheet("業務摘要")
    page_title(ws, "業務摘要", 8)

    monthly = monthly.fillna(0)
    total_sale = monthly["sales"].sum() if "sales" in monthly.columns else 0
    total_purc = monthly["purchases"].sum() if "purchases" in monthly.columns else 0
    total_gp   = monthly["gross_profit"].sum() if "gross_profit" in monthly.columns else 0
    gp_rate    = total_gp / total_sale   if total_sale else 0
    cust_cnt   = len(cust_agg)
    prod_cnt   = len(prod_agg)

    kpis = [
        ("總銷售額",   total_sale, "#,##0",  MED_BLUE),
        ("總進貨額",   total_purc, "#,##0",  DARK_BLUE),
        ("總毛利",     total_gp,   "#,##0",  GREEN if total_gp >= 0 else RED),
        ("平均毛利率", gp_rate,    "0.0%",   GREEN_SOFT),
        ("客戶數",     cust_cnt,   "#,##0",  ORANGE),
        ("產品數",     prod_cnt,   "#,##0",  "7030A0"),
    ]

    for i, (label, value, fmt, color) in enumerate(kpis):
        col = (i % 3) * 3 + 1
        row = 3 if i < 3 else 7
        ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)
        lc = ws.cell(row=row,   column=col, value=label)
        lc.font  = Font(name="Arial", size=10, color=WHITE)
        lc.fill  = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=row+1, column=col, value=value)
        vc.font  = Font(name="Arial", bold=True, size=16, color=color)
        vc.fill  = PatternFill("solid", fgColor="F7F7F7")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.number_format = fmt
        ws.row_dimensions[row+1].height = 34

    # Monthly trend table
    row = 11
    title(ws, row, 1, "月度趨勢摘要", 7)
    row += 1
    for ci, h in enumerate(["月份","銷售額","進貨額","毛利","毛利率","差額","累計銷售額"], 1):
        hdr(ws, row, ci, h, bg=MED_BLUE)
    row += 1
    cum = 0
    for i, r in monthly.sort_values("ym").iterrows():
        cum += r.get("sales", 0)
        bg = GRAY if i % 2 == 0 else WHITE
        cel(ws, row, 1, r.get("ym",""),           align="center", bg=bg)
        cel(ws, row, 2, r.get("sales",0),         "#,##0", bg=bg)
        cel(ws, row, 3, r.get("purchases",0),     "#,##0", bg=bg)
        gp = r.get("gross_profit",0)
        cel(ws, row, 4, gp,                       "#,##0", fg=(GREEN if gp>=0 else RED), bg=bg)
        mr = r.get("gp_rate", np.nan)
        cel(ws, row, 5, mr if pd.notna(mr) else 0, "0.0%", bg=bg)
        cel(ws, row, 6, r.get("balance",0),       "#,##0", bg=bg)
        cel(ws, row, 7, cum,                       "#,##0", bg=bg)
        row += 1
    border(ws, 12, row-1, 1, 7)

    widths(ws, {get_column_letter(i): 14 for i in range(1, 9)})


def sheet_rfm(wb, rfm_df):
    # RFM segmentation results: group stats + full customer detail sorted by sales
    ws = wb.create_sheet("客戶RFM分析")
    page_title(ws, "客戶RFM分群分析", 9)
    if rfm_df.empty:
        ws["A2"] = "資料不足"; return

    # Segment statistics
    grp_sum = rfm_df.groupby("segment").agg(
        customer_count=("cusno","count"),
        total_sales=("sales","sum"),
        avg_sales=("sales","mean"),
    ).reset_index().sort_values("total_sales", ascending=False)

    title(ws, 2, 1, "客群統計", 5)
    for ci, h in enumerate(["客群","客戶數","總銷售額","平均銷售額","佔比"], 1):
        hdr(ws, 3, ci, h)
    total = grp_sum["total_sales"].sum()
    for i, r in grp_sum.reset_index(drop=True).iterrows():
        row = i + 4
        bg = GRAY if i % 2 == 0 else WHITE
        cel(ws, row, 1, r["segment"],             align="left",   bg=bg)
        cel(ws, row, 2, r["customer_count"],      "#,##0",        bg=bg)
        cel(ws, row, 3, r["total_sales"],         "#,##0",        bg=bg)
        cel(ws, row, 4, r["avg_sales"],           "#,##0",        bg=bg)
        cel(ws, row, 5, r["total_sales"]/total,   "0.0%",         bg=bg)
    border(ws, 3, 3+len(grp_sum), 1, 5)

    # Detail
    row = 4 + len(grp_sum) + 2
    title(ws, row, 1, "客戶明細（按銷售額排序）", 9)
    row += 1
    show_cols = ["cusno","cusnm","segment","order_count","sales","avg_order_value","R_days","RFM_score"]
    show_cols = [c for c in show_cols if c in rfm_df.columns]
    hdrs_map  = {"cusno":"客戶編號","cusnm":"客戶名稱","segment":"客群",
                 "order_count":"訂單數","sales":"銷售額","avg_order_value":"平均客單價",
                 "R_days":"距上次購買天數","RFM_score":"RFM評分"}
    for ci, c in enumerate(show_cols, 1):
        hdr(ws, row, ci, hdrs_map.get(c, c))
    row += 1
    for i, r in rfm_df.sort_values("sales", ascending=False).reset_index(drop=True).iterrows():
        bg = GRAY if i % 2 == 0 else WHITE
        for ci, c in enumerate(show_cols, 1):
            v = r.get(c,"")
            fmt = "#,##0" if c in ["sales","order_count","avg_order_value","R_days"] else None
            cel(ws, row, ci, v, fmt=fmt, bg=bg, align="left" if c in ["cusno","cusnm","segment","RFM_score"] else "right")
        row += 1
    border(ws, 4+len(grp_sum)+3, row-1, 1, len(show_cols))

    widths(ws, {"A":14,"B":22,"C":16,"D":10,"E":14,"F":14,"G":10,"H":12,"I":12})


def sheet_pareto(wb, prod_agg, cust_agg):
    # 80/20 analysis for products and customers: who/what drives most revenue?
    ws = wb.create_sheet("帕累托分析")
    page_title(ws, "帕累托80/20分析", 8)

    # Product Pareto
    title(ws, 2, 1, "產品銷售帕累托（前80%貢獻者）", 8)
    if not prod_agg.empty:
        prd = pareto_analysis(prod_agg.copy(), "sales", "prdno")
        hdrs_p = ["排名","產品編號","產品名稱","銷售額","毛利率","佔比","累計佔比","帕累托"]
        for ci, h in enumerate(hdrs_p, 1):
            hdr(ws, 3, ci, h)
        for i, r in prd.reset_index(drop=True).iterrows():
            row = i + 4
            is_top = r["pareto"] == "前80%"
            bg = "EBF3E8" if is_top else GRAY if i%2==0 else WHITE
            mr = r.get("gp_rate", np.nan)
            cel(ws, row, 1, i+1,                       align="center", bg=bg)
            cel(ws, row, 2, r.get("prdno",""),         align="center", bg=bg)
            cel(ws, row, 3, r.get("prdnm",""),         align="left",   bg=bg)
            cel(ws, row, 4, r["sales"],                "#,##0",        bg=bg)
            cel(ws, row, 5, mr if pd.notna(mr) else 0,"0.0%",         bg=bg)
            cel(ws, row, 6, r["share"],                "0.0%",         bg=bg)
            cel(ws, row, 7, r["cumulative_share"],     "0.0%",         bg=bg)
            cel(ws, row, 8, r["pareto"],               align="center", bg=bg)
        border(ws, 3, 3+len(prd), 1, 8)
        prd_end = 3 + len(prd) + 2
    else:
        prd_end = 5

    # Customer Pareto
    title(ws, prd_end, 1, "客戶銷售帕累托", 8)
    if not cust_agg.empty:
        cst = pareto_analysis(cust_agg.copy(), "sales", "cusno")
        for ci, h in enumerate(["排名","客戶編號","客戶名稱","銷售額","佔比","累計佔比","帕累托"], 1):
            hdr(ws, prd_end+1, ci, h)
        for i, r in cst.reset_index(drop=True).iterrows():
            row = prd_end + 2 + i
            is_top = r["pareto"] == "前80%"
            bg = "EBF3E8" if is_top else GRAY if i%2==0 else WHITE
            cel(ws, row, 1, i+1,                      align="center", bg=bg)
            cel(ws, row, 2, r.get("cusno",""),         align="center", bg=bg)
            cel(ws, row, 3, r.get("cusnm",""),         align="left",   bg=bg)
            cel(ws, row, 4, r["sales"],                "#,##0",        bg=bg)
            cel(ws, row, 5, r["share"],                "0.0%",         bg=bg)
            cel(ws, row, 6, r["cumulative_share"],     "0.0%",         bg=bg)
            cel(ws, row, 7, r["pareto"],               align="center", bg=bg)
        border(ws, prd_end+1, prd_end+1+len(cst), 1, 7)

    widths(ws, {"A":8,"B":16,"C":26,"D":14,"E":10,"F":10,"G":10,"H":10})


def sheet_inventory_alert(wb, stock_agg):
    # Inventory risk view: status distribution + full detail with low-stock items highlighted
    ws = wb.create_sheet("庫存警示")
    page_title(ws, "庫存警示分析", 8)
    if stock_agg.empty:
        ws["A2"] = "資料不足"; return

    # Status summary
    if "stock_status" in stock_agg.columns:
        stat = stock_agg["stock_status"].value_counts().reset_index()
        stat.columns = ["status","item_count"]
        title(ws, 2, 1, "庫存狀態摘要", 3)
        for ci, h in enumerate(["狀態","品項數","佔比"], 1):
            hdr(ws, 3, ci, h)
        total = stat["item_count"].sum()
        for i, r in stat.iterrows():
            row = i + 4
            cel(ws, row, 1, r["status"],            align="left")
            cel(ws, row, 2, r["item_count"],        "#,##0")
            cel(ws, row, 3, r["item_count"]/total,  "0.0%")
        border(ws, 3, 3+len(stat), 1, 3)

    # Detail
    row_start = 5 + (len(stock_agg["stock_status"].unique()) if "stock_status" in stock_agg.columns else 0)
    title(ws, row_start, 1, "庫存明細（低庫存優先）", 8)
    show_cols = [c for c in ["prdno","prdnm","qty","safeqty","stock_status","stock_value","lastin","lastout"] if c in stock_agg.columns]
    hdrs_map  = {"prdno":"產品編號","prdnm":"產品名稱","qty":"現有庫存",
                 "safeqty":"安全庫存","stock_status":"庫存狀態","stock_value":"庫存價值",
                 "lastin":"最後進貨","lastout":"最後出貨"}
    for ci, c in enumerate(show_cols, 1):
        hdr(ws, row_start+1, ci, hdrs_map.get(c, c))
    for i, r in stock_agg.reset_index(drop=True).iterrows():
        row = row_start + 2 + i
        state = r.get("stock_status","")
        bg = "FFCCCC" if "Low" in str(state) or "Zero" in str(state) else (GRAY if i%2==0 else WHITE)
        for ci, c in enumerate(show_cols, 1):
            v = r.get(c,"")
            fmt = "#,##0" if c in ["qty","safeqty","stock_value"] else None
            cel(ws, row, ci, v, fmt=fmt, bg=bg, align="left" if c in ["prdno","prdnm","stock_status"] else "right")
    border(ws, row_start+1, row_start+1+len(stock_agg), 1, len(show_cols))

    widths(ws, {"A":16,"B":28,"C":12,"D":12,"E":14,"F":14,"G":14,"H":14})


def sheet_gross_margin(wb, prod_agg):
    # Margin analysis: GP tier distribution + actionable suggestions + detail tables
    ws = wb.create_sheet("毛利分析")
    page_title(ws, "產品毛利分析", 8)
    if prod_agg.empty:
        ws["A2"] = "資料不足"; return

    df = prod_agg.copy()
    df["gp_rate"] = pd.to_numeric(df["gp_rate"], errors="coerce").fillna(0)

    # GP rate distribution
    bins = [-np.inf, 0, 0.1, 0.2, 0.3, np.inf]
    labels = ["虧損(<0%)", "偏低(0-10%)", "中等(10-20%)", "良好(20-30%)", "優秀(>30%)"]
    df["gp_tier"] = pd.cut(df["gp_rate"], bins=bins, labels=labels)
    dist = df.groupby("gp_tier", observed=True).agg(item_count=("prdno","count"), sales=("sales","sum")).reset_index()

    title(ws, 2, 1, "毛利率分佈", 4)
    for ci, h in enumerate(["毛利級別","品項數","銷售佔比","建議"], 1):
        hdr(ws, 3, ci, h)
    total_sale = df["sales"].sum()
    suggestions = {
        "虧損(<0%)":    "檢討定價或停售",
        "偏低(0-10%)":  "考慮提價或降成本",
        "中等(10-20%)": "持續監控",
        "良好(20-30%)": "維持現狀",
        "優秀(>30%)":   "加強推廣",
    }
    for i, r in dist.iterrows():
        row = i + 4
        bg = GRAY if i % 2 == 0 else WHITE
        cel(ws, row, 1, str(r["gp_tier"]),                          align="left",   bg=bg)
        cel(ws, row, 2, r["item_count"],                            "#,##0",        bg=bg)
        cel(ws, row, 3, r["sales"]/total_sale if total_sale else 0, "0.0%",         bg=bg)
        cel(ws, row, 4, suggestions.get(str(r["gp_tier"]),""),      align="left",   bg=bg)
    border(ws, 3, 3+len(dist), 1, 4)

    # High/low GP product details
    row = 5 + len(dist) + 1
    for label, subset in [("低毛利產品（<10%，需關注）", df[df["gp_rate"] < 0.1].sort_values("sales", ascending=False).head(20)),
                          ("高毛利產品（>30%，加強推廣）", df[df["gp_rate"] > 0.3].sort_values("sales", ascending=False).head(20))]:
        title(ws, row, 1, label, 6)
        row += 1
        for ci, h in enumerate(["產品編號","產品名稱","銷售額","成本","毛利","毛利率"], 1):
            hdr(ws, row, ci, h)
        row += 1
        for i, r in subset.reset_index(drop=True).iterrows():
            bg = GRAY if i % 2 == 0 else WHITE
            cel(ws, row, 1, r.get("prdno",""),       align="center", bg=bg)
            cel(ws, row, 2, r.get("prdnm",""),       align="left",   bg=bg)
            cel(ws, row, 3, r["sales"],              "#,##0",        bg=bg)
            cel(ws, row, 4, r["cost"],               "#,##0",        bg=bg)
            cel(ws, row, 5, r["gross_profit"],       "#,##0", fg=(GREEN if r["gross_profit"]>=0 else RED), bg=bg)
            cel(ws, row, 6, r["gp_rate"],            "0.0%",  fg=(GREEN_SOFT if r["gp_rate"]>=0.2 else RED), bg=bg)
            row += 1
        border(ws, row - len(subset) - 1, row-1, 1, 6)
        row += 2

    widths(ws, {"A":16,"B":28,"C":14,"D":14,"E":14,"F":12})



# ══════════════════════════════════════════════════════════
# Key charts for company presentation
# ══════════════════════════════════════════════════════════

def sheet_charts(wb, monthly, cust_agg, prod_agg):
    """Standalone page: monthly revenue trend + top customers + top products"""
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("重要圖表")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:P1")
    c = ws["A1"]
    c.value = "銷售數據重要圖表"
    c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Chart 1: Monthly revenue trend line chart ─────────
    monthly_s = monthly.sort_values("ym").reset_index(drop=True) if not monthly.empty else pd.DataFrame()
    data_col_start = 20

    if not monthly_s.empty and "sales" in monthly_s.columns:
        ws.cell(row=2, column=data_col_start, value="月份")
        ws.cell(row=2, column=data_col_start+1, value="銷售額")
        ws.cell(row=2, column=data_col_start+2, value="進貨額")
        for i, r in monthly_s.iterrows():
            ws.cell(row=i+3, column=data_col_start,   value=str(r.get("ym","")))
            ws.cell(row=i+3, column=data_col_start+1, value=float(r.get("sales",0)))
            ws.cell(row=i+3, column=data_col_start+2, value=float(r.get("purchases",0)))
        n = len(monthly_s)

        chart1 = LineChart()
        chart1.title = "月度營收趨勢"
        chart1.style = 10
        chart1.height = 14
        chart1.width  = 26
        chart1.y_axis.title = "金額"
        chart1.x_axis.title = "月份"

        data_ref = Reference(ws, min_col=data_col_start+1, max_col=data_col_start+2,
                             min_row=2, max_row=2+n)
        chart1.add_data(data_ref, titles_from_data=True)
        cats = Reference(ws, min_col=data_col_start, min_row=3, max_row=2+n)
        chart1.set_categories(cats)

        chart1.series[0].graphicalProperties.line.solidFill = "2E75B6"
        chart1.series[0].graphicalProperties.line.width = 20000
        if len(chart1.series) > 1:
            chart1.series[1].graphicalProperties.line.solidFill = "ED7D31"
            chart1.series[1].graphicalProperties.line.width = 20000

        ws.add_chart(chart1, "A3")

    # ── Chart 2: TOP 10 customer bar chart ────────────────
    top_cust = cust_agg.sort_values("sales", ascending=False).head(10).reset_index(drop=True) if not cust_agg.empty else pd.DataFrame()
    data_col2 = data_col_start + 5

    if not top_cust.empty:
        ws.cell(row=2, column=data_col2,   value="客戶")
        ws.cell(row=2, column=data_col2+1, value="銷售額")
        name_col = "cusnm" if "cusnm" in top_cust.columns else "cusno"
        for i, r in top_cust.iterrows():
            ws.cell(row=i+3, column=data_col2,   value=str(r.get(name_col, r.get("cusno","")))[:12])
            ws.cell(row=i+3, column=data_col2+1, value=float(r.get("sales",0)))
        n2 = len(top_cust)

        chart2 = BarChart()
        chart2.type   = "bar"
        chart2.style  = 10
        chart2.title  = "前10大客戶銷售額"
        chart2.height = 14
        chart2.width  = 20
        chart2.y_axis.title = "客戶"
        chart2.x_axis.title = "銷售額"

        data_ref2 = Reference(ws, min_col=data_col2+1, max_col=data_col2+1,
                              min_row=2, max_row=2+n2)
        chart2.add_data(data_ref2, titles_from_data=True)
        cats2 = Reference(ws, min_col=data_col2, min_row=3, max_row=2+n2)
        chart2.set_categories(cats2)
        chart2.series[0].graphicalProperties.solidFill = "2E75B6"

        ws.add_chart(chart2, "A22")

    # ── Chart 3: TOP 10 product bar chart ─────────────────
    top_prod = prod_agg.sort_values("sales", ascending=False).head(10).reset_index(drop=True) if not prod_agg.empty else pd.DataFrame()
    data_col3 = data_col_start + 8

    if not top_prod.empty:
        ws.cell(row=2, column=data_col3,   value="產品")
        ws.cell(row=2, column=data_col3+1, value="銷售額")
        name_col3 = "prdnm" if "prdnm" in top_prod.columns else "prdno"
        for i, r in top_prod.iterrows():
            ws.cell(row=i+3, column=data_col3,   value=str(r.get(name_col3, r.get("prdno","")))[:12])
            ws.cell(row=i+3, column=data_col3+1, value=float(r.get("sales",0)))
        n3 = len(top_prod)

        chart3 = BarChart()
        chart3.type   = "bar"
        chart3.style  = 10
        chart3.title  = "前10大產品銷售額"
        chart3.height = 14
        chart3.width  = 20
        chart3.y_axis.title = "產品"
        chart3.x_axis.title = "銷售額"

        data_ref3 = Reference(ws, min_col=data_col3+1, max_col=data_col3+1,
                              min_row=2, max_row=2+n3)
        chart3.add_data(data_ref3, titles_from_data=True)
        cats3 = Reference(ws, min_col=data_col3, min_row=3, max_row=2+n3)
        chart3.set_categories(cats3)
        chart3.series[0].graphicalProperties.solidFill = "70AD47"

        ws.add_chart(chart3, "L22")

# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════

def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = input("請輸入彙總資料夾路徑 (step2 輸出):\n> ").strip()

    src = Path(folder)
    if not src.exists():
        print(f"找不到資料夾: {folder}")
        input("按 Enter 離開..."); return

    out_path = src.parent / "business_analysis_report.xlsx"
    print(f"\nSource: {src}")
    print(f"Output: {out_path}\n{'='*50}")

    monthly   = load(src, "monthly.csv")
    cust_agg  = load(src, "cust_agg.csv")
    prod_agg  = load(src, "prod_agg.csv")
    fact_agg  = load(src, "fact_agg.csv")
    stock_agg = load(src, "stock_agg.csv")

    # Numeric column conversion
    for df, cols in [
        (monthly,  ["sales","purchases","gross_profit","gp_rate","balance","sales_detail"]),
        (cust_agg, ["sales","order_count","avg_order_value","transaction_months"]),
        (prod_agg, ["sales","cost","gross_profit","gp_rate","sales_qty"]),
        (stock_agg,["qty","safeqty","stock_value"]),
    ]:
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    print("RFM 分析中...")
    rfm_df = rfm_analysis(cust_agg)

    print("建立報表中...")
    wb = Workbook()
    wb.remove(wb.active)

    sheet_charts(wb, monthly, cust_agg, prod_agg)
    sheet_summary(wb, monthly, cust_agg, prod_agg, stock_agg)
    sheet_rfm(wb, rfm_df)
    sheet_pareto(wb, prod_agg, cust_agg)
    sheet_gross_margin(wb, prod_agg)
    sheet_inventory_alert(wb, stock_agg)

    wb.save(out_path)
    print(f"\n報表完成: {out_path}")
    print("\n工作表:")
    for name in [ws.title for ws in wb.worksheets]:
        print(f"   {name}")
    input("\n按 Enter 離開...")

if __name__ == "__main__":
    main()
