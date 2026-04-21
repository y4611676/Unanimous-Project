"""
Step 6: Quarterly Executive Summary → Excel Dashboard

One-page boss-friendly summary optimised for a single-quarter data window.
Four sections:
  ① 本季整體表現    — 本季營收 + 季增率（QoQ）
  ② 最重要的發現    — 1–3 項自動偵測的關鍵洞察
  ③ 需要老闆決定的事 — 決策清單（非系統能自動解決的）
  ④ 下季預測        — 未來3個月銷售預測數字 + 折線圖

Reads same aggregated/ folder as step3/step5.
Outputs:  quarterly_executive_summary.xlsx  (in the folder's parent)
"""

import sys
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import importlib.util

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.page import PageMargins


# ── Colour palette (mirrors step3) ──────────────────────────────────────────
DARK_BLUE    = "1F3864"
MED_BLUE     = "2E75B6"
LIGHT_BLUE   = "BDD7EE"
TEAL         = "1B4F72"
GREEN        = "375623"
GREEN_SOFT   = "70AD47"
GREEN_LIGHT  = "EBF3E8"
RED          = "C00000"
ORANGE       = "ED7D31"
ORANGE_LIGHT = "FFF0E8"
GRAY         = "F2F2F2"
WHITE        = "FFFFFF"
BLACK        = "000000"
GOLD         = "FFE4B5"


# ══════════════════════════════════════════════════════════════════════════════
# Data loaders
# ══════════════════════════════════════════════════════════════════════════════

def load(folder, fname):
    """Read CSV if it exists; return empty DataFrame silently otherwise."""
    p = Path(folder) / fname
    if not p.exists():
        return pd.DataFrame()
    return pd.read_csv(p, encoding="utf-8-sig")


def _import_step5():
    """Dynamically load step5 so we can reuse sales_forecast() without circular imports."""
    spec = importlib.util.spec_from_file_location(
        "step5_advanced", Path(__file__).parent / "step5_advanced.py"
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


# ══════════════════════════════════════════════════════════════════════════════
# Data-crunching helpers
# ══════════════════════════════════════════════════════════════════════════════

def _ym_to_quarter(ym_str):
    """'2024-03' or '202403'  →  '2024-Q1'"""
    s = str(ym_str).strip().replace("-", "")
    year, month = s[:4], int(s[4:6])
    return f"{year}-Q{(month - 1) // 3 + 1}"


def quarterly_revenue(monthly):
    """Aggregate monthly.csv into [quarter, sales] sorted ascending."""
    if monthly.empty or "sales" not in monthly.columns:
        return pd.DataFrame(columns=["quarter", "sales"])
    df = monthly.copy()
    df["sales"]   = pd.to_numeric(df["sales"], errors="coerce").fillna(0)
    df["quarter"] = df["ym"].astype(str).apply(_ym_to_quarter)
    q = df.groupby("quarter", sort=True)["sales"].sum().reset_index()
    return q.sort_values("quarter").reset_index(drop=True)


def current_quarter_stats(monthly):
    """Return (label, this_q_sales, prev_q_sales, pct_change)."""
    q = quarterly_revenue(monthly)
    if q.empty:
        return "N/A", 0.0, 0.0, 0.0
    this_lbl  = str(q.iloc[-1]["quarter"])
    this_q    = float(q.iloc[-1]["sales"])
    prev_q    = float(q.iloc[-2]["sales"]) if len(q) >= 2 else 0.0
    pct       = (this_q - prev_q) / prev_q if prev_q else 0.0
    return this_lbl, this_q, prev_q, pct


def build_findings(monthly, cust_agg, prod_agg, stock_agg):
    """
    Auto-detect up to 5 findings from the aggregated data.
    Each finding is a dict with keys: title, body, action (action may be None).
    Caller picks the top 3 to display.
    """
    findings = []

    # ── Finding 1: Inventory risk ─────────────────────────────────────────────
    if not stock_agg.empty and "stock_status" in stock_agg.columns:
        bad   = int(stock_agg["stock_status"].isin(["零庫存", "低庫存"]).sum())
        total = len(stock_agg)
        if total > 0 and bad / total >= 0.10:
            findings.append({
                "title":  "庫存警示",
                "body":   (
                    f"共 {bad} 個品項（占全部 SKU 的 {bad/total:.0%}）"
                    f"處於零庫存或低庫存狀態，可能影響出貨能力。"
                ),
                "action": "滯銷品是否促銷出清？低庫存品是否立即補貨？",
            })

    # ── Finding 2: Loss-making products ──────────────────────────────────────
    if not prod_agg.empty and "gp_rate" in prod_agg.columns:
        pa = prod_agg.copy()
        pa["gp_rate"] = pd.to_numeric(pa["gp_rate"], errors="coerce").fillna(0)
        pa["sales"]   = pd.to_numeric(pa.get("sales", pd.Series(dtype=float)), errors="coerce").fillna(0)
        loss    = pa[pa["gp_rate"] < 0]
        total_s = float(pa["sales"].sum())
        if len(loss) > 0 and total_s > 0:
            loss_pct = float(loss["sales"].sum()) / total_s
            findings.append({
                "title":  "虧損品項",
                "body":   (
                    f"{len(loss)} 項產品毛利為負，"
                    f"佔本期總銷售額 {loss_pct:.1%}，建議檢討定價或成本結構。"
                ),
                "action": "虧損品項是否停售或重新定價？",
            })

    # ── Finding 3: Customer concentration risk ────────────────────────────────
    if not cust_agg.empty and "sales" in cust_agg.columns:
        ca = cust_agg.copy()
        ca["sales"] = pd.to_numeric(ca["sales"], errors="coerce").fillna(0)
        total_s  = float(ca["sales"].sum())
        if total_s > 0:
            top1     = ca.nlargest(1, "sales").iloc[0]
            top1_pct = float(top1["sales"]) / total_s
            top3_pct = float(ca.nlargest(3, "sales")["sales"].sum()) / total_s
            if top1_pct >= 0.30:
                nm = str(top1.get("cusnm", top1.get("cusno", "最大客戶")))
                findings.append({
                    "title":  "客戶集中度偏高",
                    "body":   (
                        f"最大客戶「{nm}」佔總銷售額 {top1_pct:.0%}，"
                        f"前3大客戶合計 {top3_pct:.0%}，存在單一客戶依賴風險。"
                    ),
                    "action": "是否評估客戶集中度風險，並啟動新客戶開發計畫？",
                })

    # ── Finding 4: Short-term revenue trend ──────────────────────────────────
    if not monthly.empty and "sales" in monthly.columns:
        df = monthly.copy()
        df["sales"] = pd.to_numeric(df["sales"], errors="coerce").fillna(0)
        df = df.sort_values("ym").tail(6).reset_index(drop=True)
        if len(df) >= 4:
            h1 = float(df.head(3)["sales"].mean())
            h2 = float(df.tail(3)["sales"].mean())
            if h1 > 0:
                t = (h2 - h1) / h1
                if abs(t) >= 0.10:
                    up   = t > 0
                    desc = "成長" if up else "下降"
                    note = "持續維持、可規劃進攻" if up else "需找出原因、及早因應"
                    findings.append({
                        "title":  f"近期銷售{'上升' if up else '下滑'}趨勢",
                        "body":   (
                            f"最近3個月月均銷售較前3個月{desc} {abs(t):.0%}，{note}。"
                        ),
                        "action": None,
                    })

    # ── Finding 5: High-volume but thin-margin products ───────────────────────
    if not prod_agg.empty and "sales" in prod_agg.columns and "gp_rate" in prod_agg.columns:
        pa2 = prod_agg.copy()
        pa2["sales"]  = pd.to_numeric(pa2["sales"],  errors="coerce").fillna(0)
        pa2["gp_rate"]= pd.to_numeric(pa2["gp_rate"],errors="coerce").fillna(0)
        top10 = pa2.nlargest(10, "sales")
        thin  = top10[top10["gp_rate"] < 0.10]
        if len(thin) >= 2:
            findings.append({
                "title":  "熱銷品毛利偏低",
                "body":   (
                    f"前10大熱銷品中有 {len(thin)} 項毛利率低於10%，"
                    f"量大但利薄，盈利品質需提升。"
                ),
                "action": "熱銷低毛利品是否可提價或降低採購成本？",
            })

    return findings[:5]


def build_decisions(findings):
    """Collect all decision items from findings + always-present strategic one."""
    decisions = []
    for f in findings:
        if f.get("action"):
            decisions.append({"category": f["title"], "question": f["action"]})
    decisions.append({
        "category": "下季規劃",
        "question": "根據下季預測數字，銷售目標與資源配置（人力/庫存/行銷預算）是否需要調整？",
    })
    return decisions[:5]


# ══════════════════════════════════════════════════════════════════════════════
# Excel low-level helpers
# ══════════════════════════════════════════════════════════════════════════════

def _c(ws, row, col, val=None, bold=False, size=10,
       fg=BLACK, bg=None, align="left", fmt=None, wrap=False):
    """Set a single cell's value + style."""
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def _m(ws, r1, r2, c1, c2):
    """Merge cells helper."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _sec_header(ws, row, text, bg=MED_BLUE, span=9):
    """Full-width section title bar (cols 1–span)."""
    _m(ws, row, row, 1, span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _thin_border(ws, r1, r2, c1, c2, color="CCCCCC"):
    side = Side(style="thin", color=color)
    brd  = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row:
            cell.border = brd


# ══════════════════════════════════════════════════════════════════════════════
# Executive summary sheet writer
# ══════════════════════════════════════════════════════════════════════════════

def sheet_executive_summary(wb, monthly, cust_agg, prod_agg, stock_agg, s5):
    ws = wb.create_sheet("季度經營摘要")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    # Column widths: A(left margin) B-I(content) J(right margin)
    for col, w in {1: 2, 2: 15, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 2}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ──────────────────────────────────────────────────────────────────────────
    # Banner (rows 1–2)
    # ──────────────────────────────────────────────────────────────────────────
    _m(ws, 1, 1, 1, 10)
    banner = ws.cell(row=1, column=1, value="季度經營摘要儀表板")
    banner.font      = Font(name="Arial", bold=True, size=18, color=WHITE)
    banner.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    _m(ws, 2, 2, 1, 10)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    sub = ws.cell(row=2, column=1,
                  value=f"產生時間：{ts}　　｜　　本報告由系統自動彙整，供決策參考")
    sub.font      = Font(name="Arial", size=9, color="BBBBBB")
    sub.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8  # spacer

    # ──────────────────────────────────────────────────────────────────────────
    # Section ① — 這季整體表現 (rows 4–8)
    # ──────────────────────────────────────────────────────────────────────────
    _sec_header(ws, 4, "①  這季整體表現", bg=MED_BLUE)

    q_label, this_q, prev_q, pct = current_quarter_stats(monthly)
    arrow     = "▲" if pct >= 0 else "▼"
    pct_color = GREEN_SOFT if pct >= 0 else RED

    # Card A: 本季營收 (cols 2–5)
    _m(ws, 5, 5, 2, 5)
    ca_h = ws.cell(row=5, column=2, value="本季營收")
    ca_h.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    ca_h.fill      = PatternFill("solid", fgColor=MED_BLUE)
    ca_h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 20

    _m(ws, 6, 6, 2, 5)
    ca_v = ws.cell(row=6, column=2, value=this_q)
    ca_v.font         = Font(name="Arial", bold=True, size=22, color=MED_BLUE)
    ca_v.fill         = PatternFill("solid", fgColor="EEF4FB")
    ca_v.alignment    = Alignment(horizontal="center", vertical="center")
    ca_v.number_format = "#,##0"
    ws.row_dimensions[6].height = 42

    _m(ws, 7, 7, 2, 5)
    ca_s = ws.cell(row=7, column=2, value=f"期間：{q_label}")
    ca_s.font      = Font(name="Arial", size=9, color="888888")
    ca_s.fill      = PatternFill("solid", fgColor="EEF4FB")
    ca_s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 14

    # Card B: 與上季相比 (cols 6–9)
    _m(ws, 5, 5, 6, 9)
    cb_h = ws.cell(row=5, column=6, value="與上季相比")
    cb_h.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    cb_h.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    cb_h.alignment = Alignment(horizontal="center", vertical="center")

    _m(ws, 6, 6, 6, 9)
    cb_v = ws.cell(row=6, column=6, value=f"{arrow}  {abs(pct):.1%}")
    cb_v.font      = Font(name="Arial", bold=True, size=22, color=pct_color)
    cb_v.fill      = PatternFill("solid", fgColor="F7F7F7")
    cb_v.alignment = Alignment(horizontal="center", vertical="center")

    _m(ws, 7, 7, 6, 9)
    cb_s = ws.cell(row=7, column=6, value=f"上季營收：{prev_q:,.0f}")
    cb_s.font      = Font(name="Arial", size=9, color="888888")
    cb_s.fill      = PatternFill("solid", fgColor="F7F7F7")
    cb_s.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[8].height = 10  # spacer

    # ──────────────────────────────────────────────────────────────────────────
    # Section ② — 最重要的發現 (dynamic rows, starting at 9)
    # ──────────────────────────────────────────────────────────────────────────
    _sec_header(ws, 9, "②  最重要的發現", bg=TEAL)
    row = 10

    findings = build_findings(monthly, cust_agg, prod_agg, stock_agg)
    if not findings:
        _m(ws, row, row, 2, 9)
        _c(ws, row, 2, "資料完整後將自動產生重要發現。",
           fg="888888", size=10)
        ws.row_dimensions[row].height = 20
        row += 2
    else:
        for idx, f in enumerate(findings[:3]):
            # Finding title row
            _m(ws, row, row, 2, 9)
            tc = ws.cell(row=row, column=2, value=f"  {idx + 1}.  {f['title']}")
            tc.font      = Font(name="Arial", bold=True, size=10, color=DARK_BLUE)
            tc.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
            tc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

            # Finding body row
            _m(ws, row, row, 2, 9)
            bc = ws.cell(row=row, column=2, value=f"      {f['body']}")
            bc.font      = Font(name="Arial", size=10, color="222222")
            bc.fill      = PatternFill("solid", fgColor="F5FAFF")
            bc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 22
            row += 1

            # Micro spacer between findings
            ws.row_dimensions[row].height = 5
            row += 1

    ws.row_dimensions[row].height = 8  # spacer before next section
    row += 1

    # ──────────────────────────────────────────────────────────────────────────
    # Section ③ — 需要老闆決定的事 (dynamic rows)
    # ──────────────────────────────────────────────────────────────────────────
    _sec_header(ws, row, "③  需要老闆決定的事", bg=ORANGE)
    row += 1

    decisions = build_decisions(findings)
    for i, d in enumerate(decisions):
        # Category tag (cols 2–3)
        _m(ws, row, row, 2, 3)
        lc = ws.cell(row=row, column=2, value=d["category"])
        lc.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
        lc.fill      = PatternFill("solid", fgColor=ORANGE)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        # Decision question (cols 4–9)
        _m(ws, row, row, 4, 9)
        qc = ws.cell(row=row, column=4, value=f"□   {d['question']}")
        qc.font      = Font(name="Arial", size=10, color=DARK_BLUE)
        qc.fill      = PatternFill("solid", fgColor=GOLD if i % 2 == 0 else "FFF8E7")
        qc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.row_dimensions[row].height = 10  # spacer
    row += 1

    # ──────────────────────────────────────────────────────────────────────────
    # Section ④ — 下季預測 (dynamic rows)
    # ──────────────────────────────────────────────────────────────────────────
    _sec_header(ws, row, "④  下季預測（未來3個月）", bg=GREEN)
    row += 1

    forecast_df = pd.DataFrame()
    try:
        mn = monthly.copy()
        if "sales" in mn.columns:
            mn["sales"] = pd.to_numeric(mn["sales"], errors="coerce").fillna(0)
        forecast_df = s5.sales_forecast(mn, horizon=3)
    except Exception:
        pass

    if forecast_df.empty:
        _m(ws, row, row, 2, 9)
        _c(ws, row, 2,
           "需至少6個月歷史資料才能產生預測（請先確認 monthly.csv 有足夠月份）。",
           fg="888888", size=10)
        ws.row_dimensions[row].height = 22
        row += 2
    else:
        fcast_only = (forecast_df[forecast_df["type"] == "預測"]
                      .tail(3).reset_index(drop=True))

        # ── Three forecast month cards ─────────────────────────────────────
        card_cols = [(2, 3), (4, 5), (6, 7)]

        for ci, frow in fcast_only.iterrows():
            cs, ce = card_cols[ci] if ci < len(card_cols) else (8, 9)
            # Month label
            _m(ws, row, row, cs, ce)
            mh = ws.cell(row=row, column=cs, value=str(frow["ym"]))
            mh.font      = Font(name="Arial", bold=True, size=9, color=WHITE)
            mh.fill      = PatternFill("solid", fgColor=GREEN)
            mh.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 16

            # Value
            _m(ws, row + 1, row + 2, cs, ce)
            fv = ws.cell(row=row + 1, column=cs,
                         value=round(float(frow["sales"]), 0))
            fv.font          = Font(name="Arial", bold=True, size=16, color=GREEN)
            fv.fill          = PatternFill("solid", fgColor=GREEN_LIGHT)
            fv.alignment     = Alignment(horizontal="center", vertical="center")
            fv.number_format = "#,##0"
            ws.row_dimensions[row + 1].height = 30
            ws.row_dimensions[row + 2].height = 30

        row += 4

        # Spacer before chart
        ws.row_dimensions[row].height = 6
        row += 1

        # ── Write chart data to off-screen columns (col 12+) ───────────────
        # Show last 12 actuals + 3 forecast for a meaningful trend view
        DATA_COL  = 12
        chart_src = forecast_df.tail(15).reset_index(drop=True)

        ws.cell(row=2, column=DATA_COL,     value="月份")
        ws.cell(row=2, column=DATA_COL + 1, value="銷售額")
        for j, r in chart_src.iterrows():
            ws.cell(row=j + 3, column=DATA_COL,     value=str(r["ym"]))
            ws.cell(row=j + 3, column=DATA_COL + 1, value=round(float(r["sales"]), 0))

        n = len(chart_src)
        chart = LineChart()
        chart.title   = "歷史銷售 + 未來3個月預測"
        chart.height  = 12
        chart.width   = 24
        chart.style   = 10
        chart.y_axis.title = "金額"

        data_ref = Reference(ws, min_col=DATA_COL + 1, max_col=DATA_COL + 1,
                             min_row=2, max_row=2 + n)
        chart.add_data(data_ref, titles_from_data=True)
        cats = Reference(ws, min_col=DATA_COL, min_row=3, max_row=2 + n)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.line.solidFill = GREEN_SOFT
        chart.series[0].graphicalProperties.line.width     = 20000

        ws.add_chart(chart, f"B{row}")
        row += 18

    # ── Thin border around the whole content area ──────────────────────────
    content_end = max(row - 1, 9)
    _thin_border(ws, 4, content_end, 1, 10)

    # ── Print / PDF layout — lock everything so the sheet never jumps ─────────
    ws.page_setup.orientation  = "portrait"
    ws.page_setup.paperSize    = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.2, footer=0.2
    )
    ws.print_area = f"A1:{get_column_letter(10)}{content_end}"
    ws.oddFooter.center.text = "第 &P 頁，共 &N 頁"

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) > 1:
        folder = sys.argv[1]
    else:
        folder = input("請輸入彙總資料夾路徑 (step2 輸出):\n> ").strip()

    src = Path(folder)
    if not src.exists():
        print(f"找不到資料夾: {folder}")
        input("按 Enter 離開..."); return

    out_path = src.parent / "quarterly_executive_summary.xlsx"
    print(f"\nSource : {src}")
    print(f"Output : {out_path}\n{'=' * 50}")

    monthly   = load(src, "monthly.csv")
    cust_agg  = load(src, "cust_agg.csv")
    prod_agg  = load(src, "prod_agg.csv")
    stock_agg = load(src, "stock_agg.csv")

    # Numeric conversion — same pattern as step3
    for df, cols in [
        (monthly,  ["sales", "purchases", "gross_profit", "gp_rate", "balance"]),
        (cust_agg, ["sales", "order_count", "avg_order_value"]),
        (prod_agg, ["sales", "cost", "gross_profit", "gp_rate"]),
        (stock_agg,["qty", "safeqty", "stock_value"]),
    ]:
        for c in cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Load step5 for sales_forecast(); fall back to empty stub if unavailable
    print("載入 step5 銷售預測模組...")
    try:
        s5 = _import_step5()
    except Exception as e:
        print(f"  警告：無法載入 step5（{e}），預測區段將略過。")

        class _S5Stub:
            @staticmethod
            def sales_forecast(*args, **kwargs):
                return pd.DataFrame()

        s5 = _S5Stub()

    print("建立摘要報表中...")
    wb = Workbook()
    wb.remove(wb.active)
    sheet_executive_summary(wb, monthly, cust_agg, prod_agg, stock_agg, s5)

    wb.save(out_path)
    print(f"\n摘要報表完成：{out_path}")
    for ws in wb.worksheets:
        print(f"   {ws.title}")

    # ── PDF export via Excel COM (Windows only, requires pywin32) ─────────────
    pdf_path = out_path.with_suffix(".pdf")
    try:
        import win32com.client as win32
        print("\nPDF 匯出中（透過 Excel COM）...")
        excel_app = win32.Dispatch("Excel.Application")
        excel_app.Visible        = False
        excel_app.DisplayAlerts  = False
        wb_com = excel_app.Workbooks.Open(str(out_path.resolve()))
        wb_com.ExportAsFixedFormat(
            Type               = 0,     # xlTypePDF
            Filename           = str(pdf_path.resolve()),
            Quality            = 0,     # xlQualityStandard
            IncludeDocProperties = True,
            IgnorePrintAreas   = False,
            OpenAfterPublish   = False,
        )
        wb_com.Close(False)
        excel_app.Quit()
        print(f"PDF 完成：{pdf_path}")
    except ImportError:
        print("  提示：執行 pip install pywin32 後，下次即可自動輸出同名 PDF。")
    except Exception as e:
        print(f"  PDF 轉換失敗（請確認 Microsoft Excel 已安裝）：{e}")

    input("\n按 Enter 離開...")


if __name__ == "__main__":
    main()
