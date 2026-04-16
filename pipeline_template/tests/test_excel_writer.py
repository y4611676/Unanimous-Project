"""Tests for pipeline/excel_writer.py — report generation."""

import pandas as pd
from openpyxl import load_workbook

from pipeline.excel_writer import write_report


def _sample_stage_tables():
    return {
        "cleaned": {},
        "aggregated": {
            "monthly": pd.DataFrame({
                "ym":      ["2024-01", "2024-02"],
                "revenue": [1000, 2000],
            })
        },
        "analyzed": {
            "rfm": pd.DataFrame({
                "customer_id": ["A", "B"],
                "RFM_score":   [15, 9],
            })
        },
        "anonymized": {},
    }


def test_write_report_disabled_returns_none(tmp_path):
    cfg = {"output": {"excel": {"enabled": False}}}
    assert write_report(cfg, _sample_stage_tables(), str(tmp_path)) is None


def test_write_report_emits_xlsx_with_expected_sheets(tmp_path):
    cfg = {
        "output": {
            "excel": {
                "enabled": True,
                "filename": "out.xlsx",
                "source": "aggregated",
                "sheets": [
                    {"from": "monthly", "title": "Monthly Revenue",
                     "number_cols": ["revenue"]},
                    {"from": "rfm", "stage": "analyzed", "title": "RFM"},
                ],
            }
        }
    }
    path = write_report(cfg, _sample_stage_tables(), str(tmp_path))
    assert path is not None and path.exists()

    wb = load_workbook(path)
    assert set(wb.sheetnames) == {"Monthly Revenue", "RFM"}

    ws = wb["Monthly Revenue"]
    # Row 1 is the merged page title, row 3 is the header row, data starts at row 4.
    assert ws["A3"].value == "ym"
    assert ws["B3"].value == "revenue"
    assert ws["A4"].value == "2024-01"
    assert ws["B4"].value == 1000


def test_write_report_skips_missing_tables(tmp_path):
    cfg = {
        "output": {
            "excel": {
                "enabled": True,
                "source": "aggregated",
                "sheets": [
                    {"from": "monthly", "title": "Real"},
                    {"from": "does_not_exist", "title": "Ghost"},
                ],
            }
        }
    }
    path = write_report(cfg, _sample_stage_tables(), str(tmp_path))
    assert path is not None
    wb = load_workbook(path)
    # Only the real sheet was written.
    assert wb.sheetnames == ["Real"]
