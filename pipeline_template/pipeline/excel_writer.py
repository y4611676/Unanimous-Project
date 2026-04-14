"""Formatted Excel output layer.

Adapts the styling helpers from work/step3_analyze.py (hdr, cel, border, title,
page_title, widths) so any pipeline output can be rendered as a polished
workbook without hard-coding sheet layouts.

Config shape::

    output:
      excel:
        enabled: true
        filename: report.xlsx          # written under paths.out_dir
        source: analyzed               # which subfolder to pull CSVs from
                                       # (cleaned | aggregated | analyzed | anonymized)
        sheets:
          - {from: monthly, title: 月度營收, number_cols: [revenue], percent_cols: [毛利率]}
          - {from: customer_agg, title: 客戶排行, percent_cols: [share, cum_share]}
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore


# Colours lifted from work/step3_analyze.py:17-27 so reports look familiar.
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "BDD7EE"
WHITE = "FFFFFF"
BLACK = "000000"
GRAY = "F2F2F2"

_THIN = None  # created lazily because openpyxl may not be imported at module load


def _thin_side():
    global _THIN
    if _THIN is None:
        _THIN = Side(style="thin", color="BFBFBF")
    return _THIN


def _hdr(ws, row: int, col: int, val: Any) -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cel(ws, row: int, col: int, val: Any, fmt: str | None = None,
         align: str = "right") -> None:
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        c.number_format = fmt


def _apply_border(ws, r1: int, r2: int, c1: int, c2: int) -> None:
    side = _thin_side()
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for c in row:
            c.border = border


def _page_title(ws, text: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws["A1"]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False


def _autosize(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            *(len(str(v)) for v in df[col].head(200).tolist()),
        )
        # Roughly 1.2 chars per unit width, capped so huge strings don't blow layout.
        width = min(max(max_len + 2, 10), 32)
        ws.column_dimensions[chr(64 + i) if i <= 26 else "AA"].width = width


def _fmt_for(col: str, number_cols: list[str], percent_cols: list[str]) -> str | None:
    if col in percent_cols:
        return "0.00%"
    if col in number_cols:
        return "#,##0.00"
    return None


def _write_sheet(wb, df: pd.DataFrame, title: str,
                 number_cols: list[str], percent_cols: list[str]) -> None:
    # openpyxl balks at sheet titles >31 chars or containing certain chars.
    safe_title = "".join(ch for ch in title if ch not in r"[]:*?/\\")[:31] or "Sheet"
    ws = wb.create_sheet(title=safe_title)

    n_cols = max(len(df.columns), 1)
    _page_title(ws, title, n_cols)

    header_row = 3
    for j, col in enumerate(df.columns, start=1):
        _hdr(ws, header_row, j, col)
    ws.row_dimensions[header_row].height = 22

    for i, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = ""
            fmt = _fmt_for(col, number_cols, percent_cols)
            align = "left" if isinstance(val, str) else "right"
            _cel(ws, i, j, val, fmt=fmt, align=align)

    if len(df):
        _apply_border(ws, header_row, header_row + len(df), 1, n_cols)

    _autosize(ws, df)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def write_report(config: dict[str, Any],
                 tables_by_stage: dict[str, dict[str, pd.DataFrame]],
                 out_root: str | Path) -> Path | None:
    """Emit a formatted xlsx report based on ``config['output']['excel']``.

    ``tables_by_stage`` maps stage name (cleaned/aggregated/analyzed/anonymized)
    to its ``{table_name: DataFrame}`` dict so a sheet can pull from any stage.
    Returns the path to the written workbook, or ``None`` if disabled/unavailable.
    """
    cfg = (config.get("output") or {}).get("excel") or {}
    if not cfg.get("enabled"):
        return None
    if Workbook is None:  # pragma: no cover
        print("  ⚠️  openpyxl not installed — skipping Excel output")
        return None

    default_stage = cfg.get("source", "analyzed")
    sheets_cfg = cfg.get("sheets") or []
    if not sheets_cfg:
        print("  ⚠️  output.excel.sheets is empty — skipping Excel output")
        return None

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    wrote_any = False
    for spec in sheets_cfg:
        stage = spec.get("stage", default_stage)
        name = spec.get("from")
        if not name:
            continue
        df = tables_by_stage.get(stage, {}).get(name)
        if df is None or df.empty:
            print(f"  ⚠️  excel sheet skipped (no data): stage={stage} from={name}")
            continue
        title = spec.get("title", name)
        _write_sheet(
            wb,
            df,
            title=title,
            number_cols=spec.get("number_cols", []),
            percent_cols=spec.get("percent_cols", []),
        )
        wrote_any = True

    if not wrote_any:
        return None

    out = Path(out_root)
    out.mkdir(parents=True, exist_ok=True)
    path = out / cfg.get("filename", "report.xlsx")
    wb.save(path)
    print(f"  report → {path}")
    return path
